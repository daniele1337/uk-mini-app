from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
import datetime
import os
import pandas as pd
from io import BytesIO
import base64
import json
import requests

# Загружаем переменные окружения из .env файла
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("python-dotenv не установлен. Установите: pip install python-dotenv")

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///uk_mini_app.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Telegram Bot конфигурация
TELEGRAM_BOT_TOKEN = os.environ.get('TELEGRAM_BOT_TOKEN', '8172377647:AAE6MS5TBL-tZKBWs1A3WPECef48cl_SgnU')
TELEGRAM_BOT_API_URL = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"

CORS(app)
db = SQLAlchemy(app)

# Модели базы данных
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    telegram_id = db.Column(db.String(50), unique=True, nullable=False)
    first_name = db.Column(db.String(100), nullable=False)
    last_name = db.Column(db.String(100))
    username = db.Column(db.String(100))
    apartment = db.Column(db.String(20))
    building = db.Column(db.String(20))
    street = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    is_admin = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    
    # Убираем проблемные отношения
    # meter_readings = db.relationship('MeterReading', backref='user', lazy=True)
    # complaints = db.relationship('Complaint', backref='user', lazy=True)

class MeterReading(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    meter_type = db.Column(db.String(20), nullable=False)  # electricity, cold_water, hot_water, gas, heating
    value = db.Column(db.Float, nullable=False)
    previous_value = db.Column(db.Float)
    consumption = db.Column(db.Float)  # Разница между текущим и предыдущим показанием
    notes = db.Column(db.Text)
    is_verified = db.Column(db.Boolean, default=False)
    verified_by_name = db.Column(db.String(100))  # Имя администратора, который подтвердил
    verified_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class Complaint(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    category = db.Column(db.String(50), nullable=False)  # plumbing, electricity, cleaning, noise, other
    priority = db.Column(db.String(20), default='medium')  # low, medium, high, urgent
    status = db.Column(db.String(20), default='new')  # new, in_progress, resolved, rejected, closed
    assigned_to_name = db.Column(db.String(100))  # Имя назначенного администратора
    response = db.Column(db.Text)
    resolution_notes = db.Column(db.Text)
    estimated_completion = db.Column(db.DateTime)
    actual_completion = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    message = db.Column(db.Text, nullable=False)
    title = db.Column(db.String(200))
    target = db.Column(db.String(20), default='all')  # all, active, inactive, specific
    target_users = db.Column(db.Text)  # JSON array of user IDs
    notification_type = db.Column(db.String(20), default='info')  # info, warning, success, error
    sent_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    read_by = db.Column(db.Text)  # JSON array of user IDs who read it

class MeterType(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    code = db.Column(db.String(20), unique=True, nullable=False)
    unit = db.Column(db.String(10), nullable=False)  # кВт·ч, м³, Гкал
    description = db.Column(db.Text)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

class ComplaintCategory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    code = db.Column(db.String(20), unique=True, nullable=False)
    description = db.Column(db.Text)
    sla_hours = db.Column(db.Integer, default=24)  # Service Level Agreement в часах
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

# Создание таблиц
with app.app_context():
    db.create_all()
    
    # Инициализация базовых данных
    if not MeterType.query.first():
        meter_types = [
            MeterType(name='Электричество', code='electricity', unit='кВт·ч', description='Показания электросчетчика'),
            MeterType(name='Холодная вода', code='cold_water', unit='м³', description='Показания счетчика холодной воды'),
            MeterType(name='Горячая вода', code='hot_water', unit='м³', description='Показания счетчика горячей воды'),
            MeterType(name='Газ', code='gas', unit='м³', description='Показания газового счетчика'),
            MeterType(name='Отопление', code='heating', unit='Гкал', description='Показания счетчика отопления')
        ]
        db.session.add_all(meter_types)
        
    if not ComplaintCategory.query.first():
        categories = [
            ComplaintCategory(name='Сантехника', code='plumbing', description='Проблемы с водоснабжением, канализацией', sla_hours=4),
            ComplaintCategory(name='Электричество', code='electricity', description='Проблемы с электроснабжением', sla_hours=2),
            ComplaintCategory(name='Уборка', code='cleaning', description='Проблемы с уборкой помещений', sla_hours=24),
            ComplaintCategory(name='Шум', code='noise', description='Жалобы на шум', sla_hours=48),
            ComplaintCategory(name='Другое', code='other', description='Прочие обращения', sla_hours=72)
        ]
        db.session.add_all(categories)
    
    # Создание тестового админа
    if not User.query.filter_by(telegram_id='123456789').first():
        admin_user = User(
            telegram_id='123456789',
            first_name='Администратор',
            last_name='УК',
            username='admin_uk',
            apartment='1',
            building='1',
            street='Тестовая улица',
            phone='+7 (999) 123-45-67',
            email='admin@uk.ru',
            is_admin=True,
            is_active=True
        )
        db.session.add(admin_user)
        
    db.session.commit()

# Health check endpoint
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok', 'message': 'Server is running'})

# Функции для работы с JWT
def generate_token(user_id):
    payload = {
        'user_id': user_id,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(days=30)
    }
    return jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')

def verify_token(token):
    """Проверка JWT токена"""
    try:
        if not token:
            print("No token provided")
            return None
            
        payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
        user_id = payload.get('user_id')
        
        if not user_id:
            print("No user_id in token payload")
            return None
            
        # Проверяем, что пользователь существует
        user = User.query.get(user_id)
        if not user:
            print(f"User with id {user_id} not found")
            return None
            
        return user_id
        
    except jwt.ExpiredSignatureError:
        print("Token expired")
        return None
    except jwt.InvalidTokenError as e:
        print(f"Invalid token: {str(e)}")
        return None
    except Exception as e:
        print(f"Error verifying token: {str(e)}")
        return None

# OCR функция
def process_meter_image(image_file):
    """OCR функция для распознавания показаний счетчиков"""
    try:
        # OCR функциональность временно отключена
        return {
            'success': False,
            'error': 'OCR functionality is currently disabled.'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

# API маршруты
# Улучшенная авторизация через Telegram
@app.route('/api/auth/telegram', methods=['POST'])
def telegram_auth():
    """Авторизация через Telegram Web App"""
    try:
        data = request.get_json()
        
        # Проверяем наличие обязательных полей
        if not data or 'id' not in data:
            return jsonify({'error': 'Invalid Telegram data'}), 400
        
        telegram_id = str(data['id'])
        first_name = data.get('first_name', '')
        last_name = data.get('last_name', '')
        username = data.get('username', '')
        
        # Проверяем, существует ли пользователь
        user = User.query.filter_by(telegram_id=telegram_id).first()
        
        if not user:
            # Создаем нового пользователя
            user = User(
                telegram_id=telegram_id,
                first_name=first_name,
                last_name=last_name,
                username=username,
                is_active=True
            )
            db.session.add(user)
            db.session.commit()
        
        # Генерируем токен
        token = generate_token(user.id)
        
        return jsonify({
            'success': True,
            'token': token,
            'user': {
                'id': user.id,
                'telegram_id': user.telegram_id,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'username': user.username,
                'apartment': user.apartment,
                'building': user.building,
                'street': user.street,
                'phone': user.phone,
                'email': user.email,
                'is_admin': user.is_admin,
                'is_active': user.is_active
            }
        })
        
    except Exception as e:
        print(f"Error in telegram auth: {e}")
        return jsonify({'error': 'Authentication failed'}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    
    # В реальном приложении здесь была бы проверка email/password
    # Для демо используем простую логику
    user = User.query.filter_by(email=data['email']).first()
    
    if user:
        token = generate_token(user.id)
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'telegram_id': user.telegram_id,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'username': user.username,
                'apartment': user.apartment,
                'building': user.building,
                'street': user.street,
                'phone': user.phone,
                'email': user.email
            },
            'token': token
        })
    
    return jsonify({'success': False, 'message': 'Неверные данные'}), 401

@app.route('/api/users/stats', methods=['GET'])
def user_stats():
    # Получаем пользователя из токена
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    
    token = auth_header.split(' ')[1]
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Invalid token'}), 401
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    # Статистика пользователя
    readings_count = MeterReading.query.filter_by(user_id=user_id).count()
    complaints_count = Complaint.query.filter_by(user_id=user_id).count()
    active_complaints = Complaint.query.filter_by(user_id=user_id, status='new').count()
    
    # Последние показания
    last_reading = MeterReading.query.filter_by(user_id=user_id).order_by(MeterReading.created_at.desc()).first()
    
    stats = {
        'metersCount': readings_count,
        'complaintsCount': complaints_count,
        'activeComplaints': active_complaints,
        'lastReading': {
            'electricity': last_reading.value if last_reading and last_reading.meter_type == 'electricity' else 0,
            'water': last_reading.value if last_reading and last_reading.meter_type == 'water' else 0,
            'gas': last_reading.value if last_reading and last_reading.meter_type == 'gas' else 0,
            'date': last_reading.created_at.isoformat() if last_reading else None
        } if last_reading else None
    }
    
    return jsonify(stats)

# OCR endpoint removed - functionality disabled

@app.route('/api/meters/submit', methods=['POST'])
def submit_meter_readings():
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    
    token = auth_header.split(' ')[1]
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Invalid token'}), 401
    
    data = request.json
    readings = data['readings']
    
    try:
        for meter_type, value in readings.items():
            if value:  # Если есть значение
                reading = MeterReading(
                    user_id=user_id,
                    meter_type=meter_type,
                    value=float(value)
                )
                db.session.add(reading)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Показания успешно сохранены'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/complaints', methods=['GET'])
def get_complaints():
    """Получить обращения пользователя"""
    try:
        print("=== GET /api/complaints ===")
        auth_header = request.headers.get('Authorization')
        print(f"Auth header: {auth_header}")
        
        if not auth_header or not auth_header.startswith('Bearer '):
            print("No valid auth header")
            return jsonify({'error': 'Unauthorized'}), 401
        
        token = auth_header.split(' ')[1]
        print(f"Token: {token[:20]}...")
        
        user_id = verify_token(token)
        print(f"User ID from token: {user_id}")
        
        if not user_id:
            print("Invalid token")
            return jsonify({'error': 'Invalid token'}), 401
        
        # Проверяем, существует ли пользователь
        user = User.query.get(user_id)
        if not user:
            print(f"User {user_id} not found in database")
            return jsonify({'error': 'User not found'}), 404
        
        print(f"User found: {user.first_name} {user.last_name}")
        
        complaints = Complaint.query.filter_by(user_id=user_id).order_by(Complaint.created_at.desc()).all()
        print(f"Found {len(complaints)} complaints")
        
        complaints_data = []
        for complaint in complaints:
            complaints_data.append({
                'id': complaint.id,
                'title': complaint.title,
                'description': complaint.description,
                'category': complaint.category,
                'priority': complaint.priority,
                'status': complaint.status,
                'response': complaint.response,
                'created_at': complaint.created_at.isoformat(),
                'updated_at': complaint.updated_at.isoformat()
            })
        
        print("Returning complaints data successfully")
        return jsonify({'complaints': complaints_data})
        
    except Exception as e:
        print(f"Error getting complaints: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Internal server error'}), 500

# Тестовый маршрут для проверки
@app.route('/api/test', methods=['GET'])
def test_endpoint():
    """Тестовый маршрут для проверки работы сервера"""
    try:
        print("=== GET /api/test ===")
        
        # Проверяем подключение к базе данных
        user_count = User.query.count()
        complaint_count = Complaint.query.count()
        
        return jsonify({
            'status': 'ok',
            'message': 'Server is working',
            'database': {
                'users_count': user_count,
                'complaints_count': complaint_count
            }
        })
        
    except Exception as e:
        print(f"Error in test endpoint: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# API для типов счетчиков
@app.route('/api/meter-types', methods=['GET'])
def get_meter_types():
    """Получить все типы счетчиков"""
    meter_types = MeterType.query.filter_by(is_active=True).all()
    return jsonify([{
        'id': mt.id,
        'name': mt.name,
        'code': mt.code,
        'unit': mt.unit,
        'description': mt.description
    } for mt in meter_types])

# API для категорий обращений
@app.route('/api/complaint-categories', methods=['GET'])
def get_complaint_categories():
    """Получить все категории обращений"""
    categories = ComplaintCategory.query.filter_by(is_active=True).all()
    return jsonify([{
        'id': cat.id,
        'name': cat.name,
        'code': cat.code,
        'description': cat.description,
        'sla_hours': cat.sla_hours
    } for cat in categories])

# Улучшенный API для показаний счетчиков
@app.route('/api/meters/readings', methods=['GET'])
def get_meter_readings():
    """Получить показания счетчиков пользователя"""
    try:
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        user_id = verify_token(token)
        
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Получаем последние показания по каждому типу счетчика
        readings = db.session.query(MeterReading).filter(
            MeterReading.user_id == user_id
        ).order_by(MeterReading.created_at.desc()).all()
        
        # Группируем по типу счетчика
        readings_by_type = {}
        for reading in readings:
            if reading.meter_type not in readings_by_type:
                readings_by_type[reading.meter_type] = []
            readings_by_type[reading.meter_type].append({
                'id': reading.id,
                'value': reading.value,
                'previous_value': reading.previous_value,
                'consumption': reading.consumption,
                'notes': reading.notes,
                'is_verified': reading.is_verified,
                'created_at': reading.created_at.isoformat(),
                'updated_at': reading.updated_at.isoformat()
            })
        
        return jsonify(readings_by_type)
        
    except Exception as e:
        print(f"Error getting meter readings: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/meters/readings/<meter_type>', methods=['POST'])
def submit_meter_reading(meter_type):
    """Отправить показания конкретного счетчика"""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.get_json()
    value = data.get('value')
    notes = data.get('notes', '')
    photo_data = data.get('photo')
    
    if not value:
        return jsonify({'error': 'Value is required'}), 400
    
    # Получаем предыдущее показание
    previous_reading = MeterReading.query.filter_by(
        user_id=user_id, 
        meter_type=meter_type
    ).order_by(MeterReading.created_at.desc()).first()
    
    previous_value = previous_reading.value if previous_reading else None
    consumption = value - previous_value if previous_value else None
    
    # Сохраняем фото если есть
    photo_path = None
    if photo_data:
        try:
            # Декодируем base64
            photo_bytes = base64.b64decode(photo_data.split(',')[1])
            filename = f"meter_{user_id}_{meter_type}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
            photo_path = f"uploads/{filename}"
            
            # Создаем папку если не существует
            os.makedirs('uploads', exist_ok=True)
            
            with open(photo_path, 'wb') as f:
                f.write(photo_bytes)
        except Exception as e:
            print(f"Error saving photo: {e}")
    
    # Создаем новое показание
    reading = MeterReading(
        user_id=user_id,
        meter_type=meter_type,
        value=value,
        previous_value=previous_value,
        consumption=consumption,
        notes=notes,
        photo_path=photo_path
    )
    
    db.session.add(reading)
    db.session.commit()
    
    return jsonify({
        'id': reading.id,
        'value': reading.value,
        'previous_value': reading.previous_value,
        'consumption': reading.consumption,
        'created_at': reading.created_at.strftime('%Y-%m-%d %H:%M:%S')
    })

# Улучшенный API для обращений
@app.route('/api/complaints', methods=['POST'])
def create_complaint():
    """Создать новое обращение"""
    try:
        print("=== POST /api/complaints ===")
        
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        print(f"Token: {token[:20] if token else 'None'}...")
        
        user_id = verify_token(token)
        print(f"User ID from token: {user_id}")
        
        if not user_id:
            print("No valid user_id from token")
            return jsonify({'error': 'Unauthorized'}), 401
        
        # Проверяем, существует ли пользователь
        user = User.query.get(user_id)
        if not user:
            print(f"User {user_id} not found in database")
            return jsonify({'error': 'User not found'}), 404
        
        print(f"User found: {user.first_name} {user.last_name}")
        
        data = request.get_json()
        print(f"Request data: {data}")
        
        if not data:
            print("No data provided")
            return jsonify({'error': 'No data provided'}), 400
            
        title = data.get('title')
        description = data.get('description')
        category = data.get('category')
        priority = data.get('priority', 'medium')
        
        print(f"Title: {title}")
        print(f"Description: {description}")
        print(f"Category: {category}")
        print(f"Priority: {priority}")
        
        if not all([title, description, category]):
            print("Missing required fields")
            return jsonify({'error': 'Title, description and category are required'}), 400
        
        print("Creating complaint object...")
        complaint = Complaint(
            user_id=user_id,
            title=title,
            description=description,
            category=category,
            priority=priority
        )
        
        print("Adding complaint to database...")
        db.session.add(complaint)
        db.session.commit()
        
        print(f"Complaint created successfully with ID: {complaint.id}")
        
        return jsonify({
            'success': True,
            'complaint': {
                'id': complaint.id,
                'title': complaint.title,
                'description': complaint.description,
                'category': complaint.category,
                'priority': complaint.priority,
                'status': complaint.status,
                'created_at': complaint.created_at.isoformat(),
                'updated_at': complaint.updated_at.isoformat()
            }
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error creating complaint: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    
    token = auth_header.split(' ')[1]
    current_user_id = verify_token(token)
    
    if not current_user_id or current_user_id != user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    # Обновляем поля
    for field in ['apartment', 'building', 'street', 'phone', 'email']:
        if field in data:
            setattr(user, field, data[field])
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'user': {
            'id': user.id,
            'telegram_id': user.telegram_id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'username': user.username,
            'apartment': user.apartment,
            'building': user.building,
            'street': user.street,
            'phone': user.phone,
            'email': user.email
        }
    })

# API для получения профиля пользователя
@app.route('/api/users/profile', methods=['GET'])
def get_user_profile():
    """Получить профиль пользователя"""
    try:
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        user_id = verify_token(token)
        
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'telegram_id': user.telegram_id,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'username': user.username,
                'apartment': user.apartment,
                'building': user.building,
                'street': user.street,
                'phone': user.phone,
                'email': user.email,
                'is_admin': user.is_admin,
                'is_active': user.is_active,
                'created_at': user.created_at.isoformat() if user.created_at else None
            }
        })
        
    except Exception as e:
        print(f"Error getting user profile: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

# API для обновления профиля пользователя
@app.route('/api/users/profile', methods=['PUT'])
def update_user_profile():
    """Обновить профиль пользователя"""
    try:
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        user_id = verify_token(token)
        
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Обновляем поля профиля
        if 'first_name' in data:
            user.first_name = data['first_name']
        if 'last_name' in data:
            user.last_name = data['last_name']
        if 'apartment' in data:
            user.apartment = data['apartment']
        if 'building' in data:
            user.building = data['building']
        if 'street' in data:
            user.street = data['street']
        if 'phone' in data:
            user.phone = data['phone']
        if 'email' in data:
            user.email = data['email']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'telegram_id': user.telegram_id,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'username': user.username,
                'apartment': user.apartment,
                'building': user.building,
                'street': user.street,
                'phone': user.phone,
                'email': user.email,
                'is_admin': user.is_admin,
                'is_active': user.is_active,
                'created_at': user.created_at.isoformat() if user.created_at else None
            }
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error updating user profile: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500

# Админские маршруты
@app.route('/api/admin/stats', methods=['GET'])
def admin_stats():
    # В реальном приложении здесь была бы проверка на админа
    total_users = User.query.count()
    total_complaints = Complaint.query.count()
    total_readings = MeterReading.query.count()
    active_complaints = Complaint.query.filter_by(status='new').count()
    
    return jsonify({
        'totalUsers': total_users,
        'totalComplaints': total_complaints,
        'totalReadings': total_readings,
        'activeComplaints': active_complaints
    })

@app.route('/api/admin/complaints', methods=['GET'])
def admin_complaints():
    complaints = db.session.query(Complaint, User).join(User).order_by(Complaint.created_at.desc()).all()
    
    complaints_data = []
    for complaint, user in complaints:
        complaints_data.append({
            'id': complaint.id,
            'title': complaint.title,
            'description': complaint.description,
            'category': complaint.category,
            'status': complaint.status,
            'response': complaint.response,
            'created_at': complaint.created_at.isoformat(),
            'user_name': f"{user.first_name} {user.last_name}",
            'address': f"{user.street} {user.building} {user.apartment}"
        })
    
    return jsonify({'complaints': complaints_data})

# CRM API для администраторов
@app.route('/api/admin/complaints/<int:complaint_id>', methods=['PUT'])
def update_complaint_status(complaint_id):
    """Обновить статус обращения (CRM)"""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({'error': 'Admin access required'}), 403
    
    complaint = Complaint.query.get(complaint_id)
    if not complaint:
        return jsonify({'error': 'Complaint not found'}), 404
    
    data = request.get_json()
    status = data.get('status')
    response = data.get('response')
    resolution_notes = data.get('resolution_notes')
    assigned_to = data.get('assigned_to')
    estimated_completion = data.get('estimated_completion')
    
    if status:
        complaint.status = status
        if status == 'resolved':
            complaint.actual_completion = datetime.datetime.utcnow()
    
    if response:
        complaint.response = response
    
    if resolution_notes:
        complaint.resolution_notes = resolution_notes
    
    if assigned_to:
        complaint.assigned_to_name = assigned_to # Changed from assigned_to to assigned_to_name
    
    if estimated_completion:
        complaint.estimated_completion = datetime.datetime.strptime(estimated_completion, '%Y-%m-%d %H:%M:%S')
    
    complaint.updated_at = datetime.datetime.utcnow()
    
    db.session.commit()
    
    return jsonify({
        'id': complaint.id,
        'status': complaint.status,
        'updated_at': complaint.updated_at.strftime('%Y-%m-%d %H:%M:%S')
    })

@app.route('/api/admin/meter-readings', methods=['GET'])
def admin_meter_readings():
    """Получить все показания счетчиков для админа"""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({'error': 'Admin access required'}), 403
    
    # Получаем параметры фильтрации
    meter_type = request.args.get('meter_type')
    user_filter = request.args.get('user_id')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = db.session.query(MeterReading).join(User)
    
    if meter_type:
        query = query.filter(MeterReading.meter_type == meter_type)
    
    if user_filter:
        query = query.filter(MeterReading.user_id == user_filter)
    
    if date_from:
        query = query.filter(MeterReading.created_at >= datetime.datetime.strptime(date_from, '%Y-%m-%d'))
    
    if date_to:
        query = query.filter(MeterReading.created_at <= datetime.datetime.strptime(date_to, '%Y-%m-%d'))
    
    readings = query.order_by(MeterReading.created_at.desc()).all()
    
    result = []
    for reading in readings:
        user = User.query.get(reading.user_id)
        result.append({
            'id': reading.id,
            'user_name': f"{user.first_name} {user.last_name}",
            'user_apartment': user.apartment,
            'user_building': user.building,
            'meter_type': reading.meter_type,
            'value': reading.value,
            'previous_value': reading.previous_value,
            'consumption': reading.consumption,
            'is_verified': reading.is_verified,
            'created_at': reading.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'notes': reading.notes
        })
    
    return jsonify(result)

@app.route('/api/admin/meter-readings/<int:reading_id>/verify', methods=['POST'])
def verify_meter_reading(reading_id):
    """Подтвердить показания счетчика"""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({'error': 'Admin access required'}), 403
    
    reading = MeterReading.query.get(reading_id)
    if not reading:
        return jsonify({'error': 'Reading not found'}), 404
    
    reading.is_verified = True
    reading.verified_by_name = user.first_name + ' ' + user.last_name # Changed from verified_by to verified_by_name
    reading.verified_at = datetime.datetime.utcnow()
    
    db.session.commit()
    
    return jsonify({
        'id': reading.id,
        'is_verified': reading.is_verified,
        'verified_at': reading.verified_at.strftime('%Y-%m-%d %H:%M:%S')
    })

@app.route('/api/admin/users', methods=['GET'])
def admin_users():
    users = User.query.all()
    
    users_data = []
    for user in users:
        readings_count = MeterReading.query.filter_by(user_id=user.id).count()
        complaints_count = Complaint.query.filter_by(user_id=user.id).count()
        
        users_data.append({
            'id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'username': user.username,
            'address': f"{user.street} {user.building} {user.apartment}",
            'readings_count': readings_count,
            'complaints_count': complaints_count
        })
    
    return jsonify({'users': users_data})

# API для рассылки уведомлений
@app.route('/api/admin/notifications', methods=['POST'])
def send_notification():
    """Отправка уведомлений пользователям через Telegram"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    
    token = auth_header.split(' ')[1]
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Invalid token'}), 401
    
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({'error': 'Admin access required'}), 403
    
    data = request.json
    title = data.get('title')
    message = data.get('message')
    target = data.get('target', 'all')
    building_id = data.get('building_id')
    user_ids = data.get('user_ids', [])
    notification_type = data.get('notification_type', 'info')
    
    if not title or not message:
        return jsonify({'error': 'Title and message are required'}), 400
    
    try:
        # Определяем целевых пользователей
        target_users = []
        
        if target == 'all':
            target_users = User.query.filter_by(is_active=True).all()
        elif target == 'building':
            target_users = User.query.filter_by(building=building_id, is_active=True).all()
        elif target == 'specific':
            target_users = User.query.filter(User.id.in_(user_ids), User.is_active==True).all()
        
        # Отправляем уведомления через Telegram
        telegram_result = broadcast_telegram_notification(
            title=title,
            message=message,
            notification_type=notification_type,
            target_users=target_users
        )
        
        # Сохраняем уведомления в базе данных
        for user in target_users:
            notification = Notification(
                title=title,
                message=message,
                target='specific',
                target_users=json.dumps([user.id]),
                notification_type=notification_type
            )
            db.session.add(notification)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'telegram_sent': telegram_result['success_count'],
            'telegram_failed': telegram_result['failed_count'],
            'total_users': telegram_result['total_count'],
            'message': f'Уведомление отправлено {telegram_result["success_count"]} из {telegram_result["total_count"]} пользователей через Telegram'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/telegram/test', methods=['POST'])
def test_telegram_bot():
    """Тестирование подключения к Telegram Bot"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    
    token = auth_header.split(' ')[1]
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Invalid token'}), 401
    
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({'error': 'Admin access required'}), 403
    
    try:
        # Проверяем информацию о боте
        url = f"{TELEGRAM_BOT_API_URL}/getMe"
        response = requests.get(url, timeout=10)
        bot_info = response.json()
        
        if not bot_info.get('ok'):
            return jsonify({
                'success': False,
                'error': 'Не удалось подключиться к Telegram Bot API',
                'details': bot_info
            }), 400
        
        # Отправляем тестовое сообщение администратору
        test_message = "🤖 Тестовое сообщение от бота системы управления домом\n\nЭто сообщение подтверждает, что бот работает корректно."
        
        if user.telegram_id:
            result = send_telegram_message(user.telegram_id, test_message)
            if result and result.get('ok'):
                return jsonify({
                    'success': True,
                    'bot_info': bot_info['result'],
                    'test_message_sent': True,
                    'message': 'Бот работает корректно! Тестовое сообщение отправлено.'
                })
            else:
                return jsonify({
                    'success': False,
                    'bot_info': bot_info['result'],
                    'test_message_sent': False,
                    'error': 'Не удалось отправить тестовое сообщение',
                    'details': result
                }), 400
        else:
            return jsonify({
                'success': False,
                'bot_info': bot_info['result'],
                'test_message_sent': False,
                'error': 'У администратора не указан Telegram ID'
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Ошибка при тестировании бота: {str(e)}'
        }), 500

@app.route('/api/admin/telegram/stats', methods=['GET'])
def get_telegram_stats():
    """Получение статистики по Telegram уведомлениям"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    
    token = auth_header.split(' ')[1]
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Invalid token'}), 401
    
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({'error': 'Admin access required'}), 403
    
    try:
        # Статистика пользователей с Telegram ID
        total_users = User.query.filter_by(is_active=True).count()
        users_with_telegram = User.query.filter(
            User.is_active == True,
            User.telegram_id.isnot(None)
        ).count()
        
        # Статистика уведомлений
        total_notifications = Notification.query.count()
        recent_notifications = Notification.query.filter(
            Notification.sent_at >= datetime.datetime.utcnow() - datetime.timedelta(days=7)
        ).count()
        
        return jsonify({
            'success': True,
            'stats': {
                'total_users': total_users,
                'users_with_telegram': users_with_telegram,
                'telegram_coverage_percent': round((users_with_telegram / total_users * 100) if total_users > 0 else 0, 1),
                'total_notifications': total_notifications,
                'recent_notifications': recent_notifications
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/buildings', methods=['GET'])
def get_buildings():
    """Получить список домов"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    
    token = auth_header.split(' ')[1]
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Invalid token'}), 401
    
    user = User.query.get(user_id)
    if not user or not user.is_admin:
        return jsonify({'error': 'Admin access required'}), 403
    
    try:
        # Получаем уникальные дома из пользователей
        buildings_data = db.session.query(
            User.building,
            User.street
        ).filter(
            User.building.isnot(None),
            User.street.isnot(None)
        ).distinct().all()
        
        buildings = []
        for building, street in buildings_data:
            buildings.append({
                'id': building,
                'number': building,
                'street': street,
                'full_address': f"{street}, д. {building}"
            })
        
        return jsonify({
            'success': True,
            'buildings': buildings
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/notifications', methods=['GET'])
def get_user_notifications():
    """Получить уведомления пользователя"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    
    token = auth_header.split(' ')[1]
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Invalid token'}), 401
    
    try:
        # Получаем уведомления для пользователя
        notifications = Notification.query.filter(
            db.or_(
                Notification.target == 'all',
                Notification.target == 'specific',
                Notification.target_users.contains(str(user_id))
            )
        ).order_by(Notification.sent_at.desc()).limit(50).all()
        
        notifications_data = []
        for notification in notifications:
            notifications_data.append({
                'id': notification.id,
                'title': notification.title,
                'message': notification.message,
                'type': notification.notification_type,
                'sent_at': notification.sent_at.isoformat(),
                'read_at': notification.read_by and user_id in json.loads(notification.read_by)
            })
        
        return jsonify({
            'success': True,
            'notifications': notifications_data
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/notifications/<int:notification_id>/read', methods=['PUT'])
def mark_notification_read(notification_id):
    """Отметить уведомление как прочитанное"""
    auth_header = request.headers.get('Authorization')
    if not auth_header or not auth_header.startswith('Bearer '):
        return jsonify({'error': 'Unauthorized'}), 401
    
    token = auth_header.split(' ')[1]
    user_id = verify_token(token)
    
    if not user_id:
        return jsonify({'error': 'Invalid token'}), 401
    
    try:
        notification = Notification.query.get(notification_id)
        if not notification:
            return jsonify({'error': 'Notification not found'}), 404
        
        # Добавляем пользователя в список прочитавших
        read_by = json.loads(notification.read_by) if notification.read_by else []
        if user_id not in read_by:
            read_by.append(user_id)
            notification.read_by = json.dumps(read_by)
            db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API для экспорта данных в Excel
@app.route('/api/admin/export/<export_type>', methods=['GET'])
def export_data(export_type):
    """Экспорт данных в Excel с красивым форматированием"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        wb = Workbook()
        
        if export_type == 'complaints':
            # Экспорт обращений
            ws = wb.active
            ws.title = "Обращения"
            
            # Заголовки
            headers = ['ID', 'Пользователь', 'Адрес', 'Тема', 'Описание', 'Категория', 'Приоритет', 'Статус', 'Ответ', 'Дата создания', 'Дата обновления']
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Добавляем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Получаем данные
            complaints = Complaint.query.join(User).all()
            
            # Добавляем данные
            for row, complaint in enumerate(complaints, 2):
                ws.cell(row=row, column=1, value=complaint.id)
                ws.cell(row=row, column=2, value=f"{complaint.user.first_name} {complaint.user.last_name}")
                ws.cell(row=row, column=3, value=f"{complaint.user.street}, д. {complaint.user.building}, кв. {complaint.user.apartment}")
                ws.cell(row=row, column=4, value=complaint.title)
                ws.cell(row=row, column=5, value=complaint.description)
                ws.cell(row=row, column=6, value=complaint.category)
                ws.cell(row=row, column=7, value=complaint.priority)
                ws.cell(row=row, column=8, value=complaint.status)
                ws.cell(row=row, column=9, value=complaint.response or '')
                ws.cell(row=row, column=10, value=complaint.created_at.strftime('%d.%m.%Y %H:%M'))
                ws.cell(row=row, column=11, value=complaint.updated_at.strftime('%d.%m.%Y %H:%M'))
            
            # Автоматическая ширина столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f"complaints_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
        elif export_type == 'meter_readings':
            # Экспорт показаний счетчиков
            ws = wb.active
            ws.title = "Показания счетчиков"
            
            # Заголовки
            headers = ['ID', 'Пользователь', 'Адрес', 'Тип счетчика', 'Показания', 'Предыдущие показания', 'Расход', 'Статус', 'Примечания', 'Дата создания']
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Добавляем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Получаем данные
            readings = MeterReading.query.join(User).all()
            
            # Добавляем данные
            for row, reading in enumerate(readings, 2):
                ws.cell(row=row, column=1, value=reading.id)
                ws.cell(row=row, column=2, value=f"{reading.user.first_name} {reading.user.last_name}")
                ws.cell(row=row, column=3, value=f"{reading.user.street}, д. {reading.user.building}, кв. {reading.user.apartment}")
                ws.cell(row=row, column=4, value=reading.meter_type)
                ws.cell(row=row, column=5, value=reading.value)
                ws.cell(row=row, column=6, value=reading.previous_value or '')
                ws.cell(row=row, column=7, value=reading.consumption or '')
                ws.cell(row=row, column=8, value='Подтверждено' if reading.is_verified else 'Ожидает проверки')
                ws.cell(row=row, column=9, value=reading.notes or '')
                ws.cell(row=row, column=10, value=reading.created_at.strftime('%d.%m.%Y %H:%M'))
            
            # Автоматическая ширина столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f"meter_readings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
        elif export_type == 'users':
            # Экспорт пользователей
            ws = wb.active
            ws.title = "Пользователи"
            
            # Заголовки
            headers = ['ID', 'Имя', 'Фамилия', 'Username', 'Telegram ID', 'Адрес', 'Телефон', 'Email', 'Статус', 'Админ', 'Дата регистрации']
            
            # Стили для заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Добавляем заголовки
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Получаем данные
            users = User.query.all()
            
            # Добавляем данные
            for row, user in enumerate(users, 2):
                ws.cell(row=row, column=1, value=user.id)
                ws.cell(row=row, column=2, value=user.first_name)
                ws.cell(row=row, column=3, value=user.last_name or '')
                ws.cell(row=row, column=4, value=user.username or '')
                ws.cell(row=row, column=5, value=user.telegram_id)
                ws.cell(row=row, column=6, value=f"{user.street}, д. {user.building}, кв. {user.apartment}")
                ws.cell(row=row, column=7, value=user.phone or '')
                ws.cell(row=row, column=8, value=user.email or '')
                ws.cell(row=row, column=9, value='Активен' if user.is_active else 'Неактивен')
                ws.cell(row=row, column=10, value='Да' if user.is_admin else 'Нет')
                ws.cell(row=row, column=11, value=user.created_at.strftime('%d.%m.%Y %H:%M'))
            
            # Автоматическая ширина столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
        elif export_type == 'all':
            # Экспорт всех данных в разные листы
            # Лист 1: Обращения
            ws1 = wb.active
            ws1.title = "Обращения"
            
            headers = ['ID', 'Пользователь', 'Адрес', 'Тема', 'Описание', 'Категория', 'Приоритет', 'Статус', 'Ответ', 'Дата создания']
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            complaints = Complaint.query.join(User).all()
            for row, complaint in enumerate(complaints, 2):
                ws1.cell(row=row, column=1, value=complaint.id)
                ws1.cell(row=row, column=2, value=f"{complaint.user.first_name} {complaint.user.last_name}")
                ws1.cell(row=row, column=3, value=f"{complaint.user.street}, д. {complaint.user.building}, кв. {complaint.user.apartment}")
                ws1.cell(row=row, column=4, value=complaint.title)
                ws1.cell(row=row, column=5, value=complaint.description)
                ws1.cell(row=row, column=6, value=complaint.category)
                ws1.cell(row=row, column=7, value=complaint.priority)
                ws1.cell(row=row, column=8, value=complaint.status)
                ws1.cell(row=row, column=9, value=complaint.response or '')
                ws1.cell(row=row, column=10, value=complaint.created_at.strftime('%d.%m.%Y %H:%M'))
            
            # Лист 2: Показания счетчиков
            ws2 = wb.create_sheet("Показания счетчиков")
            
            headers = ['ID', 'Пользователь', 'Адрес', 'Тип счетчика', 'Показания', 'Предыдущие показания', 'Расход', 'Статус', 'Дата создания']
            
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = header_alignment
            
            readings = MeterReading.query.join(User).all()
            for row, reading in enumerate(readings, 2):
                ws2.cell(row=row, column=1, value=reading.id)
                ws2.cell(row=row, column=2, value=f"{reading.user.first_name} {reading.user.last_name}")
                ws2.cell(row=row, column=3, value=f"{reading.user.street}, д. {reading.user.building}, кв. {reading.user.apartment}")
                ws2.cell(row=row, column=4, value=reading.meter_type)
                ws2.cell(row=row, column=5, value=reading.value)
                ws2.cell(row=row, column=6, value=reading.previous_value or '')
                ws2.cell(row=row, column=7, value=reading.consumption or '')
                ws2.cell(row=row, column=8, value='Подтверждено' if reading.is_verified else 'Ожидает проверки')
                ws2.cell(row=row, column=9, value=reading.created_at.strftime('%d.%m.%Y %H:%M'))
            
            # Лист 3: Пользователи
            ws3 = wb.create_sheet("Пользователи")
            
            headers = ['ID', 'Имя', 'Фамилия', 'Telegram ID', 'Адрес', 'Статус', 'Дата регистрации']
            
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
                cell.alignment = header_alignment
            
            users = User.query.all()
            for row, user in enumerate(users, 2):
                ws3.cell(row=row, column=1, value=user.id)
                ws3.cell(row=row, column=2, value=user.first_name)
                ws3.cell(row=row, column=3, value=user.last_name or '')
                ws3.cell(row=row, column=4, value=user.telegram_id)
                ws3.cell(row=row, column=5, value=f"{user.street}, д. {user.building}, кв. {user.apartment}")
                ws3.cell(row=row, column=6, value='Активен' if user.is_active else 'Неактивен')
                ws3.cell(row=row, column=7, value=user.created_at.strftime('%d.%m.%Y %H:%M'))
            
            # Автоматическая ширина столбцов для всех листов
            for ws in [ws1, ws2, ws3]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f"uk_mini_app_full_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        else:
            return jsonify({'error': 'Неизвестный тип экспорта'}), 400
        
        # Сохраняем файл
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Функции для работы с Telegram Bot
def send_telegram_message(chat_id, message, parse_mode='HTML'):
    """Отправка сообщения через Telegram Bot"""
    try:
        url = f"{TELEGRAM_BOT_API_URL}/sendMessage"
        data = {
            'chat_id': chat_id,
            'text': message,
            'parse_mode': parse_mode
        }
        response = requests.post(url, data=data, timeout=10)
        return response.json()
    except Exception as e:
        print(f"Error sending Telegram message: {e}")
        return None

def send_telegram_notification(user_id, title, message, notification_type='info'):
    """Отправка уведомления пользователю через Telegram"""
    try:
        # Получаем пользователя
        user = User.query.get(user_id)
        if not user or not user.telegram_id:
            return False
        
        # Форматируем сообщение в зависимости от типа
        emoji_map = {
            'info': 'ℹ️',
            'warning': '⚠️',
            'success': '✅',
            'error': '❌'
        }
        
        emoji = emoji_map.get(notification_type, 'ℹ️')
        
        # Форматируем сообщение в HTML
        formatted_message = f"""
{emoji} <b>{title}</b>

{message}

<i>Отправлено через систему управления домом</i>
        """.strip()
        
        # Отправляем сообщение
        result = send_telegram_message(user.telegram_id, formatted_message)
        return result and result.get('ok', False)
        
    except Exception as e:
        print(f"Error sending notification to user {user_id}: {e}")
        return False

def broadcast_telegram_notification(title, message, notification_type='info', target_users=None):
    """Массовая рассылка уведомлений через Telegram"""
    if not target_users:
        target_users = User.query.filter_by(is_active=True).all()
    
    success_count = 0
    total_count = len(target_users)
    
    for user in target_users:
        if send_telegram_notification(user.id, title, message, notification_type):
            success_count += 1
    
    return {
        'success_count': success_count,
        'total_count': total_count,
        'failed_count': total_count - success_count
    }

if __name__ == '__main__':
    # Создаем таблицы в базе данных
    with app.app_context():
        # Удаляем старую базу данных и создаем новую
        import os
        db_path = 'instance/uk_mini_app.db'
        if os.path.exists(db_path):
            print("Removing old database...")
            os.remove(db_path)
            print("Old database removed")
        
        db.create_all()
        print("Database tables created successfully")
        
        # Создаем тестового пользователя, если его нет
        test_user = User.query.filter_by(telegram_id='123456789').first()
        if not test_user:
            test_user = User(
                telegram_id='123456789',
                first_name='Тестовый',
                last_name='Пользователь',
                username='test_user',
                apartment='15',
                building='3',
                street='Ленина',
                phone='+7 (999) 123-45-67',
                email='test@example.com',
                is_admin=False,
                is_active=True
            )
            db.session.add(test_user)
            db.session.commit()
            print("Test user created successfully")
        
        # Создаем тестовые типы счетчиков
        meter_types = [
            {'name': 'Электричество', 'code': 'electricity', 'unit': 'кВт·ч'},
            {'name': 'Холодная вода', 'code': 'cold_water', 'unit': 'м³'},
            {'name': 'Горячая вода', 'code': 'hot_water', 'unit': 'м³'},
            {'name': 'Газ', 'code': 'gas', 'unit': 'м³'},
            {'name': 'Отопление', 'code': 'heating', 'unit': 'Гкал'}
        ]
        
        for mt_data in meter_types:
            existing = MeterType.query.filter_by(code=mt_data['code']).first()
            if not existing:
                meter_type = MeterType(**mt_data)
                db.session.add(meter_type)
        
        # Создаем тестовые категории обращений
        complaint_categories = [
            {'name': 'Сантехника', 'code': 'plumbing', 'description': 'Проблемы с водоснабжением, канализацией'},
            {'name': 'Электрика', 'code': 'electricity', 'description': 'Проблемы с электричеством'},
            {'name': 'Отопление', 'code': 'heating', 'description': 'Проблемы с отоплением'},
            {'name': 'Уборка', 'code': 'cleaning', 'description': 'Проблемы с уборкой подъездов'},
            {'name': 'Шум', 'code': 'noise', 'description': 'Жалобы на шум'},
            {'name': 'Лифт', 'code': 'elevator', 'description': 'Проблемы с лифтом'},
            {'name': 'Ремонт', 'code': 'repair', 'description': 'Общие вопросы ремонта'},
            {'name': 'Общие вопросы', 'code': 'general', 'description': 'Другие вопросы'}
        ]
        
        for cat_data in complaint_categories:
            existing = ComplaintCategory.query.filter_by(code=cat_data['code']).first()
            if not existing:
                category = ComplaintCategory(**cat_data)
                db.session.add(category)
        
        db.session.commit()
        print("Test data initialized successfully")
    
    app.run(debug=True, host='0.0.0.0', port=8000) 